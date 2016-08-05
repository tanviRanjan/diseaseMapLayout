# -*- coding: utf-8 -*-
"""
Created on Mon May 30 11:17:18 2016

@author: tanviranjan
"""

import openpyxl
'''File to merge all the map files together'''

'''step 1: read the names of the file from the master file'''
wb = openpyxl.load_workbook('/Users/tanviranjan/Desktop/excel files/MasterFile1_(for_data_analysis)V6_cut.xlsx')
sh = wb.active

cell = 'Eosinophil' #search all files for this cell type
count = 0

for i in sh.rows[0]:
    if (i.value == 'Cell type'):
        cell_type = count
    
    if(i.value == 'File name'):
        file_name = count
        
    count = count + 1
    
fileList = list()

'''For each of the file names read, add Map_ in the front to be able to read from the folder excel files/maps/'''
for i in range(1,sh.max_row + 1):
    cell_name = sh.cell(row = i, column = cell_type + 1).value
    name_file = sh.cell(row = i, column = file_name + 1).value
    if (cell_name == cell and name_file is not None and not(name_file == 'N/A')):
        fname = 'Map_';
        fileList.append(fname + sh.cell(row = i, column = file_name + 1).value)
    
wb.save('/Users/tanviranjan/Desktop/excel files/MasterFile1_(for_data_analysis)V6_cut.xlsx')
    
'''Read all the groups alongwith the compartment information'''
wb2 = openpyxl.load_workbook('/Users/tanviranjan/Desktop/excel files/edgelist/' + cell + '.xlsx')
shw = wb2.active

rowno = 1
for i in range(0,len(fileList)):
    fname = fileList[i];
    print (fname)
    fname = '/Users/tanviranjan/Desktop/excel files/maps/'+ fname;
    
    wb = openpyxl.load_workbook(fname)
    sh = wb.active
    
    '''sh: map file (to be read from), shw: Mast cell (to write into)'''    
    
    k = 2 # Key values for reading from the file
    start = rowno # This keeps track of the row number in the toWriteIn file
    for i in range(start, start + sh.max_row):
        #print (sh.cell(row = k, column = 1).value)
        shw.cell(row = i, column = 5).value = sh.cell(row = k, column = 1).value
        #print (sh.cell(row = i, column = fro+1).value)
        shw.cell(row = i, column = 6).value = sh.cell(row = k, column = 2).value
        
        rowno = rowno + 1
        k = k + 1
    
    wb.save(fname)

# There are 3 blank spaces after each file, need to be removed    

'''---------------------------------------------------------------------------------------------------------'''
'''step 3: now add the grouping info for extracellular > cell membrane > cytoplasm > nucleus'''
shw.cell(row = rowno + 1, column = 5).value = 'nucleus'
shw.cell(row = rowno + 1, column = 6).value = 'cytoplasm'

rowno = rowno + 1

shw.cell(row = rowno + 1, column = 5).value = 'cytoplasm'
shw.cell(row = rowno + 1, column = 6).value = 'cell membrane'

rowno = rowno + 1

shw.cell(row = rowno + 1, column = 5).value = 'cell membrane'
shw.cell(row = rowno + 1, column = 6).value = 'extracellular'
'''-------------------------------------------------------------------------------------------------------------'''

'''step 4: remove duplicates''' # done in excel for now

wb2.save('/Users/tanviranjan/Desktop/excel files/edgelist/' + cell + '.xlsx')

    
