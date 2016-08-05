# -*- coding: utf-8 -*-
"""
Created on Wed May 25 10:30:40 2016

@author: tanviranjan
"""

import openpyxl
'''File to merge all the files together'''

'''step 1: read the names of the file from the master file'''
wb = openpyxl.load_workbook('/Users/tanviranjan/Desktop/excel files/MasterFile1_(for_data_analysis)V6_cut.xlsx')
sh = wb.active

cell = 'Airway epithelial cell' #search all files for this cell type
count = 0

for i in sh.rows[0]:
    if (i.value == 'Cell type'):
        cell_type = count
    
    if(i.value == 'File name'):
        file_name = count
        
    count = count + 1
    
fileList = list()
for i in range(1,sh.max_row + 1):
    if (sh.cell(row = i, column = cell_type + 1).value == cell):
        f = sh.cell(row = i, column = file_name + 1).value
        if(f is not None and not(f == 'N/A')):
            fileList.append(f)
            
print ('----------------------')
 
'''for i in range(0, len(fileList)):       
    print (fileList[i])'''
    
wb.save('/Users/tanviranjan/Desktop/excel files/MasterFile1_(for_data_analysis)V6_cut.xlsx')


# replace the '(Genes)' file by '(Interactions)' file
fileListInt = list(); 
for i in range(0, len(fileList)):
    fname = fileList[i]
    if (fname is None or 'IL-10' in fname ):  # information for the IL 10 file is not available
        # do nothing
        print ('none')
    else: 
        if ('(Genes)' in fname):
            fileListInt.append(fname.replace('(Genes)','(Interactions)'))
        else:
            fileListInt.append(fname.replace('.xls','(1).xls'))
        
for i in fileListInt:        
    print (i)

'''--------------------------------------------------------------------------------'''
'''step 2: read those files and add to one set'''
'''--------------------------------------------------------------------------------'''
row = 1;

wb2 = openpyxl.load_workbook('/Users/tanviranjan/Desktop/excel files/edgelist/' + cell + '.xlsx')

shw = wb2.active
for n in range(0, len(fileListInt)):
    f = fileListInt[n];
    fname = '/Users/tanviranjan/Desktop/excel files/xlsx files/' + str(f); #source of xlsx files (in a different folder)
    wb = openpyxl.load_workbook(fname)
    sh = wb.active

    count = 0
    for i in sh.rows[2]:
        #print (type(i.value))
        if(i.value == 'Network Object "FROM"'):
            fro = count
        if(i.value == 'Network Object "TO"'):
            to = count
        if(i.value == 'Effect'):
            effect = count
        
        count = count + 1
        
    

    '''This to be done for every file in advance
    shw.cell(row = 1,column = 1).value = 'source'
    shw.cell(row = 1,column = 2).value = 'destination'
    shw.cell(row = 1,column = 3).value = 'effect' '''

    # Copy data from the workbook
    k = 4 # Key values for reading from the file
    start = row # This keeps track of the row number in the toWriteIn file
    for i in range(start, start + sh.max_row):
        #print (sh.cell(row = k, column = fro + 1).value)
        shw.cell(row = i, column = 1).value = sh.cell(row = k, column = fro + 1).value
        #print (sh.cell(row = i, column = fro+1).value)
        shw.cell(row = i, column = 2).value = sh.cell(row = k, column = to + 1).value
    
        if(sh.cell(row = k, column = effect + 1).value == 'Activation'):
            shw.cell(row = i, column = 3).value = '+1'
        else:
            shw.cell(row = i, column = 3).value = '-1' 
        
        row = row + 1
        k = k + 1

    wb.save(fname)
# There are 3 blank spaces after each file, need to be removed    
wb2.save('/Users/tanviranjan/Desktop/excel files/edgelist/' + cell + '.xlsx')

'''step 3: remove duplicate interactions''' # for now, done in excel

'''repeat all the steps for the (genes) file''' 